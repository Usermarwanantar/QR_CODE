import streamlit as st
import qrcode
from PIL import Image, ImageDraw, ImageFont
import io
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
import pandas as pd

# Configuration de la page
st.set_page_config(
    page_title="QR Code Generator Pro", 
    page_icon="🎯", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personnalisé pour un design moderne
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
        box-shadow: 0 8px 32px rgba(0,0,0,0.1);
    }
    
    .card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        margin-bottom: 1rem;
        border: 1px solid #f0f0f0;
    }
    
    .form-card {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.1);
    }
    
    .history-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin-bottom: 1rem;
    }
    
    .success-message {
        background: linear-gradient(135deg, #56ab2f 0%, #a8e6cf 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .warning-message {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .stButton > button {
        border-radius: 25px;
        border: none;
        padding: 0.5rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.2);
    }
    
    .qr-display {
        text-align: center;
        padding: 2rem;
        background: white;
        border-radius: 15px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.1);
        margin: 2rem 0;
    }
    
    .stats-container {
        display: flex;
        justify-content: space-around;
        margin: 2rem 0;
    }
    
    .stat-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        flex: 1;
        margin: 0 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# En-tête principal avec design moderne
st.markdown("""
<div class="main-header">
    <h1>🎯 QR Code Generator Pro</h1>
    <p style="font-size: 1.2rem; margin-top: 0.5rem;">Générateur intelligent de QR Codes pour fichiers OneDrive</p>
</div>
""", unsafe_allow_html=True)

# Instructions dans une carte moderne
with st.expander("📚 Guide d'utilisation", expanded=False):
    st.markdown("""
    <div class="card">
        <h4>🚀 Étapes pour créer votre QR Code :</h4>
        <ol>
            <li><strong>Téléversez</strong> votre fichier sur <a href='https://onedrive.live.com/' target='_blank'>OneDrive</a></li>
            <li><strong>Activez le partage</strong> (Toute personne avec le lien...)</li>
            <li><strong>Copiez le lien</strong> et collez-le dans le formulaire ci-dessous</li>
        </ol>
    </div>
    """, unsafe_allow_html=True)

# Formulaire dans une carte moderne
st.markdown("""
<div class="form-card">
    <h3 style="text-align: center; margin-bottom: 2rem; color: #2c3e50;">📝 Formulaire de création</h3>
</div>
""", unsafe_allow_html=True)

# Formulaire en colonnes pour un meilleur layout
col1, col2 = st.columns(2)

with col1:
    project_name = st.text_input("🏷️ Nom du projet", placeholder="Ex: Projet Alpha")
    dtr = st.text_input("🗂️ DTR", placeholder="Ex: DTR-2024-001")
    indice = st.text_input("✏️ Indice", placeholder="Ex: IND-001")

with col2:
    shared_link = st.text_input("🔗 Lien OneDrive", placeholder="https://1drv.ms/...")
    file_type = st.selectbox("📂 Type de fichier", ["Plugmap", "WireList", "Gamme d'assemblage"])
    file_title = st.text_input("📄 Titre du document", placeholder="Ex: Plan d'assemblage")

# Fichiers
qr_image_folder = "qr_images"
os.makedirs(qr_image_folder, exist_ok=True)

# Statistiques en haut
history_file = "historique1_qr.xlsx"
if os.path.exists(history_file):
    try:
        df_history = pd.read_excel(history_file)
        total_qr = len(df_history)
        unique_projects = df_history["Nom du projet"].nunique()
        
        st.markdown("""
        <div class="stats-container">
            <div class="stat-card">
                <h3>📊 Statistiques</h3>
                <p><strong>Total QR Codes:</strong> """ + str(total_qr) + """</p>
                <p><strong>Projets uniques:</strong> """ + str(unique_projects) + """</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
    except Exception as e:
        # Si le fichier est corrompu, on l'affiche pas les stats
        st.warning("⚠️ Le fichier d'historique est corrompu. Les statistiques ne peuvent pas être affichées.")
        pass

# Bouton de génération centré
st.markdown("<br>", unsafe_allow_html=True)
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    generate_button = st.button("🚀 Générer le QR Code", use_container_width=True)

# Logique de génération
if shared_link and file_title and file_type and project_name and dtr and indice:
    if generate_button:
        # Génération QR enrichi
        qr = qrcode.make(shared_link)
        qr = qr.resize((300, 300))
        width, height = 400, 420
        qr_image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(qr_image)

        try:
            font_title = ImageFont.truetype("arial.ttf", 18)
            font_type = ImageFont.truetype("arial.ttf", 16)
        except:
            font_title = font_type = None

        type_text = f"[📂 {file_type} File]"
        bbox_type = draw.textbbox((0, 0), type_text, font=font_type)
        w_type = bbox_type[2] - bbox_type[0]
        draw.text(((width - w_type) / 2, 10), type_text, fill="black", font=font_type)

        qr_image.paste(qr, (50, 40))

        bbox_title = draw.textbbox((0, 0), file_title, font=font_title)
        w_title = bbox_title[2] - bbox_title[0]
        draw.text(((width - w_title) / 2, 360), file_title, fill="black", font=font_title)

        # Affichage du résultat dans une carte moderne
        st.markdown("""
        <div class="qr-display">
            <h3 style="color: #2c3e50; margin-bottom: 1rem;">🎉 QR Code généré avec succès !</h3>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.image(qr_image, caption="QR Code prêt à être partagé", use_container_width=True)

        # Sauvegarde image
        qr_filename = f"{file_title.replace(' ', '_')}_QR.png"
        qr_path = os.path.join(qr_image_folder, qr_filename)
        qr_image.save(qr_path)

        # Boutons d'action
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            with open(qr_path, "rb") as f:
                st.download_button(
                    label="📥 Télécharger QR Code",
                    data=f,
                    file_name=qr_filename,
                    mime="image/png",
                    use_container_width=True
                )
        
        # Création Excel si nécessaire
        if not os.path.exists(history_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Nom du projet", "DTR", "Indice", "Titre", "Type", "Lien partagé", "QR Code"])
            wb.save(history_file)

        # Chargement Excel avec gestion d'erreur
        try:
            wb = load_workbook(history_file)
            ws = wb.active
            next_row = ws.max_row + 1
        except Exception as e:
            # Si le fichier est corrompu, on le supprime et on en crée un nouveau
            st.warning("Le fichier Excel existant est corrompu. Création d'un nouveau fichier...")
            if os.path.exists(history_file):
                os.remove(history_file)
            wb = Workbook()
            ws = wb.active
            ws.append(["Nom du projet", "DTR", "Indice", "Titre", "Type", "Lien partagé", "QR Code"])
            next_row = 2

        # Données texte
        ws.cell(row=next_row, column=1).value = project_name
        ws.cell(row=next_row, column=2).value = dtr
        ws.cell(row=next_row, column=3).value = indice
        ws.cell(row=next_row, column=4).value = file_title
        ws.cell(row=next_row, column=5).value = file_type
        ws.cell(row=next_row, column=6).value = shared_link

        # Image QR centrée dans G{next_row}
        img = XLImage(qr_path)
        img.width = 100
        img.height = 100
        cell_ref = f"G{next_row}"
        ws.add_image(img, cell_ref)
        ws.row_dimensions[next_row].height = 120

        # Mise en forme entêtes
        headers = ["Nom du projet", "DTR", "Indice", "Titre", "Type", "Lien partagé", "QR Code"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Largeur des colonnes
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 18

        # Centrage du texte
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for idx, cell in enumerate(row):
                if idx == 5:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        try:
            wb.save(history_file)
            st.markdown("""
            <div class="success-message">
                ✅ QR Code enregistré avec succès dans l'historique !
            </div>
            """, unsafe_allow_html=True)
            
            # Bouton de téléchargement Excel après sauvegarde réussie
            with col2:
                if os.path.exists(history_file):
                    with open(history_file, "rb") as f:
                        st.download_button(
                            label="📊 Télécharger Excel",
                            data=f,
                            file_name=history_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
            
            st.rerun()
        except PermissionError:
            st.markdown("""
            <div class="warning-message">
                ❌ Fichier Excel ouvert. Fermez-le puis réessayez.
            </div>
            """, unsafe_allow_html=True)
else:
    if generate_button:
        st.markdown("""
        <div class="warning-message">
            🟠 Veuillez remplir tous les champs pour générer le QR Code.
        </div>
        """, unsafe_allow_html=True)

# Section Historique avec design moderne
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown("""
<div class="history-card">
    <h3 style="text-align: center; margin-bottom: 1rem;">📚 Historique des QR Codes</h3>
</div>
""", unsafe_allow_html=True)

if os.path.exists(history_file):
    try:
        df_history = pd.read_excel(history_file)
        
        # Bouton pour supprimer tout l'historique
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🗑️ Supprimer tout l'historique", type="secondary", use_container_width=True):
                st.session_state.show_confirm_delete_all = True
            
            if st.session_state.get('show_confirm_delete_all', False):
                if st.button("⚠️ Confirmer la suppression", type="primary", use_container_width=True):
                    try:
                        os.remove(history_file)
                        # Supprimer aussi les images QR
                        for filename in os.listdir(qr_image_folder):
                            if filename.endswith("_QR.png"):
                                os.remove(os.path.join(qr_image_folder, filename))
                        st.success("✅ Tout l'historique a été supprimé.")
                        st.session_state.show_confirm_delete_all = False
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Erreur lors de la suppression : {e}")
                if st.button("❌ Annuler", use_container_width=True):
                    st.session_state.show_confirm_delete_all = False
                    st.rerun()
        
        # Bouton de téléchargement Excel toujours visible
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if os.path.exists(history_file):
                with open(history_file, "rb") as f:
                    st.download_button(
                        label="📊 Télécharger l'historique Excel",
                        data=f,
                        file_name=history_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.warning("Aucun fichier d'historique disponible.")
        
        # Affichage de l'historique dans une carte moderne
        st.markdown("""
        <div class="card">
            <h4 style="margin-bottom: 1rem;">📋 Liste des QR Codes générés</h4>
        </div>
        """, unsafe_allow_html=True)
        
        # Affichage personnalisé avec images QR
        headers = ["Nom du projet", "DTR", "Indice", "Titre", "Type", "Lien partagé", "QR Code", "Action"]
        cols = st.columns([2, 1, 1, 2, 1, 3, 1, 1])
        for i, h in enumerate(headers):
            cols[i].markdown(f"**{h}**")
        
        for idx, row in df_history.iterrows():
            cols = st.columns([2, 1, 1, 2, 1, 3, 1, 1])
            cols[0].write(row.get("Nom du projet", ""))
            cols[1].write(row.get("DTR", ""))
            cols[2].write(row.get("Indice", ""))
            cols[3].write(row.get("Titre", ""))
            cols[4].write(row.get("Type", ""))
            cols[5].write(row.get("Lien partagé", ""))
            qr_filename = f"{str(row.get('Titre', '')).replace(' ', '_')}_QR.png"
            qr_path = os.path.join(qr_image_folder, qr_filename)
            if os.path.exists(qr_path):
                cols[6].image(qr_path, width=120)
            else:
                cols[6].write("(Non trouvé)")
            
            # Bouton de suppression pour cette ligne
            with cols[7]:
                if st.button(f"❌", key=f"del_{idx}", help="Supprimer cette ligne"):
                    st.session_state[f'confirm_delete_{idx}'] = True
                
                if st.session_state.get(f'confirm_delete_{idx}', False):
                    if st.button(f"⚠️ Confirmer", key=f"confirm_{idx}"):
                        try:
                            # Supprimer la ligne du DataFrame
                            df_history = df_history.drop(idx)
                            # Sauvegarder le DataFrame mis à jour
                            df_history.to_excel(history_file, index=False)
                            # Supprimer l'image QR correspondante
                            if os.path.exists(qr_path):
                                os.remove(qr_path)
                            st.success(f"✅ Ligne supprimée.")
                            st.session_state[f'confirm_delete_{idx}'] = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Erreur lors de la suppression : {e}")
                    if st.button(f"❌ Annuler", key=f"cancel_{idx}"):
                        st.session_state[f'confirm_delete_{idx}'] = False
                        st.rerun()
        
        # Sélecteur de titre et affichage du QR code correspondant
        titres = df_history["Titre"].dropna().unique().tolist()
        if titres:
            st.markdown("""
            <div class="card">
                <h4 style="margin-bottom: 1rem;">🔍 Afficher un QR Code spécifique</h4>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                selected_titre = st.selectbox("Choisissez un titre pour voir son QR code :", titres)
                if selected_titre:
                    qr_filename = f"{str(selected_titre).replace(' ', '_')}_QR.png"
                    qr_path = os.path.join(qr_image_folder, qr_filename)
                    if os.path.exists(qr_path):
                        st.image(qr_path, caption=f"QR Code pour : {selected_titre}", width=300)
                    else:
                        st.warning("QR code non trouvé pour ce titre.")
    except Exception as e:
        st.markdown("""
        <div class="warning-message">
            ❌ Le fichier d'historique est corrompu. Il sera supprimé et recréé lors de la prochaine génération de QR Code.
        </div>
        """, unsafe_allow_html=True)
        # Supprimer le fichier corrompu
        try:
            os.remove(history_file)
        except:
            pass
else:
    st.markdown("""
    <div class="card">
        <p style="text-align: center; color: #7f8c8d;">Aucun historique trouvé. Générez un QR Code pour commencer à enregistrer l'historique.</p>
    </div>
    """, unsafe_allow_html=True)
